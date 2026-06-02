def read_legacy(path):
    """
    Reads 2020-2023.xlsx.
    Each raw data tab is a separate year/tournament.
    Returns dict of {tab_name: dataframe}.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    
    # Only grab raw data tabs, skip metadata tabs
    raw_tabs = [s for s in wb.sheetnames 
                if '原始' in s or s.startswith('COA')]
    
    result = {}
    for tab in raw_tabs:
        df = pd.read_excel(path, sheet_name=tab, header=1)
        df['_source_tab'] = tab      # track where each row came from
        df['_source_file'] = path
        result[tab] = df
    return result


def read_modern(path):
    """
    Reads 2024+ single-season files.
    Each file has separate sheets: 原始数据, 对局数据, 赛后数据.
    Returns dict of {sheet_type: dataframe}.
    """
    sheets_wanted = ['原始数据', '对局数据', '赛后数据']
    
    result = {}
    for sheet in sheets_wanted:
        try:
            df = pd.read_excel(path, sheet_name=sheet, header=1)
            df['_source_file'] = path
            result[sheet] = df
        except Exception:
            # sheet doesn't exist in this file — skip silently
            pass
    return result


def parse_filename_metadata(filename):
    """
    Extracts year, league, season from modern filenames.
    e.g. '2024IVL夏季赛常规赛.xlsx' -> 
         {'year': 2024, 'league': 'IVL', 'season': '夏季赛', 'stage': '常规赛'}
    """
    import re
    name = filename.replace('.xlsx', '')
    
    year_match = re.match(r'(\d{4})', name)
    year = int(year_match.group(1)) if year_match else None
    
    league = None
    for l in ['IVL', 'IJL', 'IVS', 'COA']:
        if l in name:
            league = l
            break
    
    season = '夏季赛' if '夏季赛' in name else '秋季赛' if '秋季赛' in name else None
    stage = '常规赛' if '常规赛' in name else '季后赛' if '季后赛' in name else None
    
    return {
        'year': year, 
        'league': league, 
        'season': season, 
        'stage': stage
    }